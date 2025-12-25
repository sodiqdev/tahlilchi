from openpyxl.utils import get_column_letter
from .styles import *
from .formula import *


def build_title(ws, config, last_col_letter):
    title = (
        f"{config['tuman']}dagi {config['maktab']}ning {config['sinf']} "
        f"{config['fan']} fanidan o'tkazilgan 2025-2026 o'quv yili{config['chorak']}-chorak "
        f"{config['imtihon_nomi']} tahlili"
    )

    ws.merge_cells(f'A1:{last_col_letter}1')
    cell = ws['A1']
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = CENTER_WRAP
    ws.row_dimensions[1].height = 50


def build_header(ws, num_tasks, jami_col_letter, percent_col_letter):
    ws.merge_cells('A2:A3')
    ws['A2'].value = "№"
    ws['A2'].font = BOLD
    ws['A2'].alignment = CENTER

    ws.merge_cells('B2:B3')
    ws['B2'].value = "FISH"
    ws['B2'].font = BOLD
    ws['B2'].alignment = CENTER

    for i in range(1, num_tasks + 1):
        col_letter = get_column_letter(2 + i)
        cell = ws[f"{col_letter}2"]
        cell.value = f"{i}-topshiriq maks.ball"
        cell.font = BOLD
        cell.alignment = CENTER_WRAP

    ws[jami_col_letter + "2"] = "Jami ball"
    ws[jami_col_letter + "2"].alignment = CENTER
    ws[jami_col_letter + "2"].font = BOLD

    ws[percent_col_letter + "2"] = "%"
    ws[percent_col_letter + "2"].alignment = CENTER
    ws[percent_col_letter + "2"].font = BOLD

    ws.merge_cells(f"{percent_col_letter}2:{percent_col_letter}3")


def build_max_scores_row(ws, max_scores, num_tasks, jami_col_letter):
    for idx, score in enumerate(max_scores, start=1):
        col_letter = get_column_letter(2 + idx)
        cell = ws[f"{col_letter}3"]
        cell.value = score
        cell.alignment = CENTER
        cell.font = RED_BOLD
        cell.fill = YELLOW_FILL

    start_idx = 3
    end_idx = 2 + num_tasks
    formula = build_max_row_total_formula(start_idx, end_idx)

    cell = ws[f"{jami_col_letter}3"]
    cell.value = formula
    cell.alignment = CENTER
    cell.font = RED_BOLD
    cell.fill = YELLOW_FILL


def build_student_rows(ws, students_list, num_tasks, jami_col_letter, percent_col_letter):
    row = 4
    for idx, name in enumerate(students_list, start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=name)

        for col in range(3, 3 + num_tasks):
            ws.cell(row=row, column=col, value="")

        # Jami formula
        sum_formula = build_sum_formula(3, 2 + num_tasks, row)
        c = ws[f"{jami_col_letter}{row}"]
        c.value = sum_formula
        c.font = RED_BOLD
        c.fill = YELLOW_FILL

        # % formula
        total_col_letter = get_column_letter(num_tasks + 3)
        percent_formula = build_percentage_formula(total_col_letter, row)
        ws[f"{percent_col_letter}{row}"].value = percent_formula

        row += 1

    return row  # next "Jami" row


def build_footer(ws, jami_row, num_tasks):
    start_row = 4
    end_row = jami_row - 1

    ws.cell(jami_row, 2, "Jami")

    for col in range(3, 3 + num_tasks):
        col_letter = get_column_letter(col)
        formula = build_average_formula(col_letter, start_row, end_row)
        cell = ws.cell(row=jami_row, column=col, value=formula)
        cell.font = RED_BOLD
        cell.fill = YELLOW_FILL

    # Jami
    total_col = num_tasks + 3
    col_letter = get_column_letter(total_col)
    formula = build_average_formula(col_letter, start_row, end_row)
    c = ws.cell(row=jami_row, column=total_col, value=formula)
    c.font = RED_BOLD
    c.fill = YELLOW_FILL

    # %
    percent_col = num_tasks + 4
    col_letter = get_column_letter(percent_col)
    formula = build_average_formula(col_letter, start_row, end_row)
    c = ws.cell(row=jami_row, column=percent_col, value=formula)
    c.font = RED_BOLD
    c.fill = YELLOW_FILL

    return jami_row


def build_signatures(ws, jami_row, total_columns, config):
    signature_row = jami_row + 2
    end_col_letter = get_column_letter(min(4, total_columns))

    ws.merge_cells(f"A{signature_row}:{end_col_letter}{signature_row}")
    ws[f"A{signature_row}"] = f"O'IBDO': _____________ {config['oibdo']}"

    ws.merge_cells(f"A{signature_row+2}:{end_col_letter}{signature_row+2}")
    ws[f"A{signature_row+2}"] = f"Metod birlashma rahbari: _____________ {config['metod_rahbari']}"

    ws.merge_cells(f"A{signature_row+4}:{end_col_letter}{signature_row+4}")
    ws[f"A{signature_row+4}"] = f"Fan o'qituvchisi: _____________ {config['fan_oqituvchisi']}"
