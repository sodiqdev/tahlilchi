from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .validators import (
    validate_students_list,
    validate_num_tasks,
    validate_max_scores,
    prepare_config,
)
from .builder import (
    build_title,
    build_header,
    build_max_scores_row,
    build_student_rows,
    build_footer,
    build_signatures,
)
from .styles import *
from .file_utils import get_safe_filename, ensure_output_dir


def create_assessment_template(students_list, num_tasks, config=None,
                               max_scores=None, output_dir="."):

    # 1. Validatsiya
    validate_students_list(students_list)
    validate_num_tasks(num_tasks)
    config = prepare_config(config)
    max_scores = validate_max_scores(max_scores, num_tasks)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tahlil"

    total_columns = 2 + num_tasks + 2
    last_col_letter = get_column_letter(total_columns)
    jami_col_letter = get_column_letter(2 + num_tasks + 1)
    percent_col_letter = get_column_letter(total_columns)

    # 2. Sarlavha
    build_title(ws, config, last_col_letter)

    # 3. Header
    build_header(ws, num_tasks, jami_col_letter, percent_col_letter)

    # 4. Max scores row
    build_max_scores_row(ws, max_scores, num_tasks, jami_col_letter)

    # 5. Students
    jami_row = build_student_rows(ws, students_list, num_tasks,
                                  jami_col_letter, percent_col_letter)

    # 6. Footer (jami)
    build_footer(ws, jami_row, num_tasks)

    # 7. Signatures
    build_signatures(ws, jami_row, total_columns, config)

    # 8. Style & Borders
    for row in ws.iter_rows(min_row=2, max_row=jami_row, min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = THIN_BORDER

    # Left-align name column
    for row in ws.iter_rows(min_row=3, max_row=jami_row, min_col=1, max_col=2):
        for c in row:
            c.alignment = LEFT

    for row in ws.iter_rows(min_row=4, max_row=jami_row, min_col=3, max_col=2 + num_tasks):
        for c in row:
            c.alignment = CENTER

    for row in ws.iter_rows(min_row=4, max_row=jami_row,
                            min_col=2 + num_tasks + 1, max_col=total_columns):
        for c in row:
            c.alignment = CENTER

    # 9. Column widths
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 30

    for i in range(num_tasks):
        ws.column_dimensions[get_column_letter(3 + i)].width = 11

    ws.column_dimensions[get_column_letter(num_tasks + 3)].width = 8
    ws.column_dimensions[get_column_letter(num_tasks + 4)].width = 8

    # 10. Save
    ensure_output_dir(output_dir)
    filename = f"{config['sinf']}_{config['fan']}_{config['chorak']}_chorak.xlsx"
    filename = get_safe_filename(filename)
    file_path = f"{output_dir}/{filename}"
    wb.save(file_path)

    return {
        'file_path': file_path,
        'total_students': len(students_list),
        'total_tasks': num_tasks,
        'config': config
    }
