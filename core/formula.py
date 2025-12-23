from openpyxl.utils import get_column_letter

def build_sum_formula(start_col_idx: int, end_col_idx: int, row: int) -> str:
    start_letter = get_column_letter(start_col_idx)
    end_letter = get_column_letter(end_col_idx)
    return f"=SUM({start_letter}{row}:{end_letter}{row})"


def build_max_row_total_formula(start_col_idx: int, end_col_idx: int) -> str:
    start_letter = get_column_letter(start_col_idx)
    end_letter = get_column_letter(end_col_idx)
    return f"=SUM({start_letter}3:{end_letter}3)"


def build_percentage_formula(total_col_letter: str, row: int) -> str:
    return f"={total_col_letter}{row}/${total_col_letter}$3*100"


def build_average_formula(col_letter: str, start_row: int, end_row: int) -> str:
    return f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})"
